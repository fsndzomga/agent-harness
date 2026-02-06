"""Custom smolagents tools for GAIA file handling."""

from smolagents import tool


@tool
def read_file_tool(file_path: str) -> str:
    """Read and extract text content from a file. Supports PDF, XLSX, CSV,
    TXT, and common text formats.

    Args:
        file_path: Path to the file to read.

    Returns:
        Extracted text content from the file.
    """
    import os

    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".pdf":
            return _read_pdf(file_path)
        elif ext in (".xlsx", ".xls"):
            return _read_excel(file_path)
        elif ext == ".csv":
            return _read_csv(file_path)
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            return f"[Image file: {file_path}] — Use code to process this image."
        elif ext in (".mp3", ".wav", ".m4a", ".ogg"):
            return f"[Audio file: {file_path}] — Audio transcription not available."
        else:
            with open(file_path, "r", errors="replace") as f:
                return f.read()[:50000]
    except Exception as e:
        return f"Error reading {file_path}: {e}"


def _read_pdf(path: str) -> str:
    import pymupdf

    doc = pymupdf.open(path)
    text = ""
    for page in doc:
        text += page.get_text() + "\n"
    doc.close()
    return text[:50000]


def _read_excel(path: str) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            result.append("\t".join(str(c) if c is not None else "" for c in row))
    wb.close()
    return "\n".join(result)[:50000]


def _read_csv(path: str) -> str:
    with open(path, "r", errors="replace") as f:
        return f.read()[:50000]
